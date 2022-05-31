from argparse import Namespace
from data.models import is_valid, Game, LegislativeSession, LegislativeOutcome, Party, Player, PresidentAction, PresidentActionType, Role, WinReason
from datetime import datetime
from openpyxl import load_workbook

import config
import data.repository as repo


def _read_spreadsheet(file: str) -> list[list]:
    workbook = load_workbook(filename=file, data_only=True)
    worksheet = workbook[config.WORKSHEET_NAME]
    if worksheet.max_column != config.SPREADSHEET_NUM_COLS:
        raise ValueError(f"Wrong number of columns in spreadsheet. Expected {config.SPREADSHEET_NUM_COLS} but received {worksheet.max_column}.")
    data = []
    for row in range(1, worksheet.max_row + 1):
        row_data = [worksheet.cell(row=row, column=col).value for col in range(1, worksheet.max_column + 1)]
        data.append(row_data)
    return data


def _is_empty(row: list) -> bool:
    return all([x is None for x in row])


def _is_date(date_str: str) -> bool:
    if not isinstance(date_str, str):
        return False
    try:
        datetime.strptime(date_str, "%Y-%m-%d")
        return True
    except ValueError:
        return False


def _is_date_row(row: list[str]) -> bool:
    return _is_date(row[0]) and _is_empty(row[1:])


def _is_header_row(row: list[str]) -> bool:
    return row == config.SPREADSHEET_HEADER


def _get_value(row: list, column: str):
    index = config.SPREADSHEET_HEADER.index(column)
    return row[index]


def _parse_data(data: list[list]) -> tuple[list[Game], list[Player], list[LegislativeSession], list[PresidentAction]]:
    games = []
    players = []
    leg_sessions = []
    pres_actions = []
    new_game = False
    for row in data:
        if _is_empty(row):
            continue
        elif _is_date_row(row):
            current_date = datetime.strptime(row[0], "%Y-%m-%d")
        elif _is_header_row(row):
            new_game = True
        else:
            # Save the new game
            if new_game:
                game_id = _get_value(row, "game")
                winning_team = Party(_get_value(row, "winning_team"))
                win_reason = WinReason(_get_value(row, "win_reason"))
                games.append(Game(game_id, current_date, winning_team, win_reason))
                new_game = False
            # Add player if present
            player_name = _get_value(row, "player")
            player_role = _get_value(row, "role")
            if player_name is not None or player_role is not None:
                players.append(Player(game_id, player_name, Role(player_role)))
            # Add legislative session if present
            round_num = _get_value(row, "round")
            if round_num is not None:
                pres_name = _get_value(row, "president")
                chan_name = _get_value(row, "chancellor")
                outcome = LegislativeOutcome(_get_value(row, "outcome"))
                top_deck_raw = _get_value(row, "top_deck")
                top_deck = None if top_deck_raw is None else Party(top_deck_raw)
                pres_get_claim = _get_value(row, "pres_get")
                pres_give_claim = _get_value(row, "pres_give")
                chan_get_claim = _get_value(row, "chan_get")
                veto_attempt_raw = _get_value(row, "veto_attempt")
                veto_attempt = False if not veto_attempt_raw else veto_attempt_raw
                last_round_raw = _get_value(row, "last_round")
                last_round = False if not last_round_raw else last_round_raw
                pres_get_actual = _get_value(row, "pres_get_actual")
                chan_get_actual = _get_value(row, "chan_get_actual")
                leg_sessions.append(LegislativeSession(game_id, round_num, pres_name, chan_name, outcome, top_deck, pres_get_claim, pres_give_claim, chan_get_claim, pres_get_actual, chan_get_actual, veto_attempt, last_round))
                # Add president action if present
                pres_action = _get_value(row, "pres_action")
                if pres_action is not None:
                    action = PresidentActionType(pres_action)
                    target_name = _get_value(row, "target")
                    num_lib = _get_value(row, "num_lib")
                    accuse = _get_value(row, "accuse")
                    pres_actions.append(PresidentAction(game_id, round_num, action, target_name, num_lib, accuse))
    return games, players, leg_sessions, pres_actions


def main(args: Namespace) -> None:
    file = args.file
    print(f"Starting to import data from '{file}'.")
    data = _read_spreadsheet(file)
    games, players, leg_sessions, pres_actions = _parse_data(data)
    if is_valid(games, players, leg_sessions, pres_actions):
        repo.clear_all()
        print(f"Saving {len(games)} games.")
        for g in games:
            repo.save_game(g)
        print(f"Saving {len(players)} players.")
        for p in players:
            repo.save_player(p)
        print(f"Saving {len(leg_sessions)} legislative sessions.")
        for ls in leg_sessions:
            repo.save_leg_session(ls)
        print(f"Saving {len(pres_actions)} president actions.")
        for a in pres_actions:
            repo.save_pres_action(a)
    print("Import complete.")
